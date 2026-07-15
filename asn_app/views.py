from django.conf import settings
import logging
import re
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.timezone import now
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Q, Count
from .models import ASN, SuratPerintahTugas, KopSurat, SuratSantunanKorpri, NotaDinas, HariLibur, SuratCuti, SisaCuti, Siswa, SuratKeterangan, SuratResmi, SPTJM, SPMT, FotoKegiatan, SuratUmum, SuratPanggilanSiswa, SiswaKeluar, SuratRekomendasiStudiLanjut, SuratKP4, AnggotaKeluargaKP4, SuratUndanganSiswa, PesertaNotaDinas, SuratDispensasi, PesertaDispensasi, SuratUsulan, PesertaSuratUsulan, StSatyalancana
from .forms import ASNForm, SPTForm, KopSuratForm, SuratSantunanKorpriForm, NotaDinasForm, HariLiburForm, SuratCutiForm, SisaCutiForm, SiswaForm, SuratKeteranganForm, SuratResmiForm, SPTJMForm, SPMTForm, FotoKegiatanForm, SuratUmumForm, SuratPanggilanSiswaForm, SiswaKeluarForm, SuratRekomendasiStudiLanjutForm, SuratKP4Form, AnggotaKeluargaKP4FormSet, SuratUndanganSiswaForm, PesertaNotaDinasForm, PesertaNotaDinasCRUDForm, PesertaNotaDinasFormSet, SuratDispensasiForm, PesertaDispensasiFormSet, SuratUsulanForm, PesertaSuratUsulanForm, PesertaSuratUsulanCRUDForm, PesertaSuratUsulanFormSet, StSatyalancanaForm, DRHSatyalancanaForm
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy, reverse
import base64
import os



def asn_list(request):
    """Menampilkan daftar semua ASN"""
    query = request.GET.get('q')
    if query:
        asn_queryset = ASN.objects.filter(nama__icontains=query).order_by('nama')
    else:
        asn_queryset = ASN.objects.all().order_by('nama')

    paginator = Paginator(asn_queryset, 10)  # 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'asn_app/asn_list.html', {'page_obj': page_obj, 'query': query})

def asn_detail(request, pk):
    """Menampilkan detail ASN"""
    asn = get_object_or_404(ASN, pk=pk)
    current_year = now().year

    # Get available years from surat_cuti for this ASN
    from django.db.models import Min, Max
    year_range = SuratCuti.objects.filter(pegawai=asn).aggregate(
        min_year=Min('tanggal_awal__year'),
        max_year=Max('tanggal_awal__year')
    )

    # Generate list of years: current year (N), N-1, N-2
    years = []
    for i in range(3):
        years.append(current_year - i)

    # Get SisaCuti for this ASN
    try:
        sisa_cuti = SisaCuti.objects.get(pegawai=asn)
    except SisaCuti.DoesNotExist:
        sisa_cuti = None

    # Calculate used leave days and remaining balance for each year (n, n-1, n-2)
    leave_data = {}
    total_initial = 0  # Total initial allocation (before deduct)
    total_remaining_after_deduct = 0  # Total remaining after deduct

    for year in years:
        # Get all surat cuti for this year
        surat_cuti_qs = SuratCuti.objects.filter(
            pegawai=asn,
            tanggal_awal__year=year
        )

        # Calculate total used days
        used_days = sum(sc.calculate_effective_leave_days() for sc in surat_cuti_qs)

        # Get initial allocation and remaining balance from SisaCuti
        if sisa_cuti:
            if year == current_year:
                initial = sisa_cuti.initial_tahun_n
                remaining = sisa_cuti.sisa_tahun_n
            elif year == current_year - 1:
                initial = sisa_cuti.initial_tahun_n_1
                remaining = sisa_cuti.sisa_tahun_n_1
            else:  # current_year - 2
                initial = sisa_cuti.initial_tahun_n_2
                remaining = sisa_cuti.sisa_tahun_n_2
        else:
            initial = 0
            remaining = 0

        total_initial += initial
        total_remaining_after_deduct += remaining

        leave_data[year] = {
            'used': used_days,
            'initial': initial,  # Initial allocation (before deduct)
            'remaining': remaining  # Remaining after deduct
        }

    return render(request, 'asn_app/asn_detail.html', {
        'asn': asn,
        'years': years,
        'leave_data': leave_data,
        'current_year': current_year,
        'total_initial': total_initial,
        'total_remaining_after_deduct': total_remaining_after_deduct
    })

def asn_leave_history(request, pk, year, leave_type):
    """Menampilkan riwayat cuti ASN untuk tahun tertentu"""
    asn = get_object_or_404(ASN, pk=pk)
    current_year = now().year

    # Get surat cuti for the specified year
    surat_cuti_list = SuratCuti.objects.filter(
        pegawai=asn,
        tanggal_awal__year=year
    ).order_by('-tanggal_surat')

    # Calculate used days for each surat cuti
    leave_details = []
    for sc in surat_cuti_list:
        leave_details.append({
            'surat_cuti': sc,
            'days': sc.calculate_effective_leave_days()
        })

    total_days = sum(item['days'] for item in leave_details)

    # Get initial allocation and remaining balance from SisaCuti
    try:
        sisa_cuti = SisaCuti.objects.get(pegawai=asn)
        if year == current_year:
            initial = sisa_cuti.initial_tahun_n
            remaining = sisa_cuti.sisa_tahun_n
        elif year == current_year - 1:
            initial = sisa_cuti.initial_tahun_n_1
            remaining = sisa_cuti.sisa_tahun_n_1
        else:  # current_year - 2
            initial = sisa_cuti.initial_tahun_n_2
            remaining = sisa_cuti.sisa_tahun_n_2
    except SisaCuti.DoesNotExist:
        initial = 0
        remaining = 0

    context = {
        'asn': asn,
        'year': year,
        'leave_type': leave_type,
        'leave_details': leave_details,
        'total_days': total_days,
        'initial': initial,
        'remaining': remaining
    }

    return render(request, 'asn_app/asn_leave_history.html', context)

def asn_create(request):
    """Membuat ASN baru"""
    if request.method == 'POST':
        form = ASNForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('asn_list')
    else:
        form = ASNForm()
    return render(request, 'asn_app/asn_form.html', {'form': form, 'title': 'Tambah ASN'})

def asn_update(request, pk):
    """Mengupdate data ASN"""
    asn = get_object_or_404(ASN, pk=pk)
    if request.method == 'POST':
        form = ASNForm(request.POST, request.FILES, instance=asn)
        if form.is_valid():
            form.save()
            return redirect('asn_detail', pk=asn.pk)
    else:
        form = ASNForm(instance=asn)
    return render(request, 'asn_app/asn_form.html', {'form': form, 'title': 'Edit ASN'})

def asn_delete(request, pk):
    """Menghapus data ASN"""
    asn = get_object_or_404(ASN, pk=pk)
    if request.method == 'POST':
        asn.delete()
        return redirect('asn_list')
    return render(request, 'asn_app/asn_confirm_delete.html', {'asn': asn})

def asn_export_pdf(request, pk):
    """Export profil ASN ke PDF"""
    from weasyprint import HTML

    asn = get_object_or_404(ASN, pk=pk)

    # Render template HTML dengan current time
    html_string = render_to_string('asn_app/asn_pdf_template.html', {
        'asn': asn,
        'current_time': now(),
        'request': request,
    })

    # Create PDF
    try:
        html = HTML(string=html_string)
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)


    # Create HTTP response dgn PDF
    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=profil_{asn.nip}_{asn.nama}.pdf'

    return response

def asn_print_view(request, pk):
    """Tampilan khusus untuk print"""
    asn = get_object_or_404(ASN, pk=pk)
    return render(request, 'asn_app/asn_print.html', {
        'asn': asn,
        'current_time': now()
    })

# SPT Views
def spt_list(request):
    """Menampilkan daftar semua SPT"""
    spt_list = SuratPerintahTugas.objects.all().order_by('-created_at')
    return render(request, 'asn_app/spt_list.html', {'spt_list': spt_list})

def spt_detail(request, pk):
    """Menampilkan detail SPT"""
    spt = get_object_or_404(SuratPerintahTugas, pk=pk)
    fotos = spt.foto_kegiatan.all()
    return render(request, 'asn_app/spt_detail.html', {'spt': spt, 'fotos': fotos})


def spt_laporan(request, pk):
    """Menampilkan laporan SPT"""
    spt = get_object_or_404(SuratPerintahTugas, pk=pk)
    fotos = spt.foto_kegiatan.all()

    # Group photos into sets of 4
    grouped_fotos = []
    for i in range(0, len(fotos), 4):
        grouped_fotos.append(fotos[i:i + 4])

    return render(request, 'asn_app/spt_laporan.html', {'spt': spt, 'grouped_fotos': grouped_fotos})


def upload_foto_kegiatan(request, spt_pk):
    spt = get_object_or_404(SuratPerintahTugas, pk=spt_pk)
    if request.method == 'POST':
        # Form ini sekarang hanya berisi 'keterangan'
        form = FotoKegiatanForm(request.POST)
        if form.is_valid():
            files = request.FILES.getlist('foto')

            # Cek jika tidak ada file yang di-upload
            if not files:
                messages.error(request, 'Tidak ada file foto yang dipilih untuk diupload.')
                # Render kembali halaman dengan form yang ada
                return render(request, 'asn_app/upload_foto_kegiatan.html', {'form': form, 'spt': spt})

            for f in files:
                instance = FotoKegiatan(
                    spt=spt,
                    foto=f,
                    keterangan=form.cleaned_data['keterangan']
                )
                instance.save()

            messages.success(request, f'{len(files)} foto berhasil diupload.')
            return redirect('spt_detail', pk=spt.pk)
    else:
        form = FotoKegiatanForm()
    return render(request, 'asn_app/upload_foto_kegiatan.html', {'form': form, 'spt': spt})


def delete_foto_kegiatan(request, foto_pk):
    foto = get_object_or_404(FotoKegiatan, pk=foto_pk)
    spt_pk = foto.spt.pk
    if request.method == 'POST':
        foto.delete()
        return redirect('spt_detail', pk=spt_pk)
    return render(request, 'asn_app/delete_foto_kegiatan.html', {'foto': foto})


def cetak_foto_kegiatan_pdf(request, spt_pk):
    """Export activity photos to PDF"""
    import os
    import base64
    from weasyprint import HTML

    spt = get_object_or_404(SuratPerintahTugas, pk=spt_pk)
    fotos = spt.foto_kegiatan.all()

    # Encode images to base64
    encoded_fotos = []
    for foto in fotos:
        if foto.foto and os.path.exists(foto.foto.path):
            try:
                with open(foto.foto.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = foto.foto.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    encoded_foto = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
                    encoded_fotos.append({
                        'base64': encoded_foto,
                        'keterangan': foto.keterangan
                    })
            except Exception as e:
                logging.error(f"Error encoding image to base64: {e}")

    # Group photos into sets of four for display on A4 page (2x2 grid)
    grouped_fotos = []
    for i in range(0, len(encoded_fotos), 4):
        grouped_fotos.append(encoded_fotos[i:i + 4])

    # Render template HTML
    html_string = render_to_string('asn_app/cetak_foto_kegiatan_pdf.html', {
        'spt': spt,
        'grouped_fotos': grouped_fotos,
    })

    # Create PDF
    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    # Create HTTP response dgn PDF
    response = HttpResponse(result, content_type='application/pdf')
    safe_nomor = re.sub(r'[^\w\-]', '_', spt.nomor_spt)
    response['Content-Disposition'] = f'attachment; filename="foto_kegiatan_{safe_nomor}.pdf"'

    return response


def spt_create(request):
    """Membuat SPT baru"""
    if request.method == 'POST':
        form = SPTForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('spt_list')
    else:
        form = SPTForm()
    return render(request, 'asn_app/spt_form.html', {'form': form, 'title': 'Tambah SPT'})

def spt_update(request, pk):
    """Mengupdate data SPT"""
    spt = get_object_or_404(SuratPerintahTugas, pk=pk)
    if request.method == 'POST':
        form = SPTForm(request.POST, instance=spt)
        if form.is_valid():
            form.save()
            return redirect('spt_detail', pk=spt.pk)
    else:
        form = SPTForm(instance=spt)
    return render(request, 'asn_app/spt_form.html', {'form': form, 'title': 'Edit SPT'})

def spt_delete(request, pk):
    """Menghapus data SPT"""
    spt = get_object_or_404(SuratPerintahTugas, pk=pk)
    if request.method == 'POST':
        spt.delete()
        return redirect('spt_list')
    return render(request, 'asn_app/spt_confirm_delete.html', {'spt': spt})

def spt_export_pdf(request, pk):
    """Export SPT ke PDF"""
    import os
    import base64
    from weasyprint import HTML

    spt = get_object_or_404(SuratPerintahTugas, pk=pk)

    # Debug logging
    logger = logging.getLogger(__name__)
    logger.info(f"spt: {spt}")
    logger.info(f"spt.kop_surat: {spt.kop_surat}")

    kop_surat_base64 = None
    if spt.kop_surat and spt.kop_surat.gambar:
        logger.info(f"spt.kop_surat.gambar: {spt.kop_surat.gambar}")
        logger.info(f"spt.kop_surat.gambar.url: {spt.kop_surat.gambar.url}")
        logger.info(f"spt.kop_surat.gambar.path: {spt.kop_surat.gambar.path}")

        # Cek apakah file benar-benar ada
        if os.path.exists(spt.kop_surat.gambar.path):
            logger.info("File gambar KOP SURAT ADA di filesystem")
            # Encode gambar ke base64
            try:
                with open(spt.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = spt.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
                    logger.info("Gambar berhasil di-encode ke base64")
            except Exception as e:
                logger.error(f"Error encoding image to base64: {e}")
        else:
            logger.error("File gambar KOP SURAT TIDAK ADA di filesystem")

    # Choose template based on participant count
    template_name = 'asn_app/spt_pdf_template_large.html' if spt.peserta.all().count() > 3 else 'asn_app/spt_pdf_template.html'

    # Render template HTML
    html_string = render_to_string(template_name, {
        'spt': spt,
        'current_time': now(),
        'request': request,
        'MEDIA_URL': '/media/',  # Tambahkan MEDIA_URL secara manual
        'kop_surat_base64': kop_surat_base64,
    })

    # Create PDF
    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logger.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    # Create HTTP response dgn PDF
    response = HttpResponse(result, content_type='application/pdf')
    safe_nomor = re.sub(r'[^\w\-]', '_', spt.nomor_spt)
    safe_nama = re.sub(r'[^\w\-]', '_', spt.nama_kegiatan)
    response['Content-Disposition'] = f'attachment; filename="spt_{safe_nomor}_{safe_nama}.pdf"'

    return response

def spt_print_view(request, pk):
    """Tampilan khusus untuk print SPT"""
    spt = get_object_or_404(SuratPerintahTugas, pk=pk)
    return render(request, 'asn_app/spt_print.html', {
        'spt': spt,
        'current_time': now()
    })

def spt_export_pdf_large(request, pk):
    """Export SPT ke PDF dengan daftar peserta lengkap"""
    import os
    import base64
    from weasyprint import HTML

    spt = get_object_or_404(SuratPerintahTugas, pk=pk)

    logger = logging.getLogger(__name__)

    kop_surat_base64 = None
    if spt.kop_surat and spt.kop_surat.gambar:
        if os.path.exists(spt.kop_surat.gambar.path):
            try:
                with open(spt.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = spt.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logger.error(f"Error encoding image to base64 for large PDF: {e}")

    html_string = render_to_string('asn_app/spt_pdf_template_large.html', {
        'spt': spt,
        'current_time': now(),
        'request': request,
        'MEDIA_URL': '/media/',
        'kop_surat_base64': kop_surat_base64,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logger.error(f"Error writing large PDF: {e}")
        return HttpResponse(f"Error writing large PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    safe_nomor = re.sub(r'[^\w\-]', '_', spt.nomor_spt)
    safe_nama = re.sub(r'[^\w\-]', '_', spt.nama_kegiatan)
    response['Content-Disposition'] = f'attachment; filename="spt_lengkap_{safe_nomor}_{safe_nama}.pdf"'

    return response


# Kop Surat Views
def kop_surat_list(request):
    """Menampilkan daftar semua Kop Surat"""
    kop_surat_list = KopSurat.objects.all().order_by('-created_at')
    return render(request, 'asn_app/kop_surat_list.html', {'kop_surat_list': kop_surat_list})

def kop_surat_detail(request, pk):
    """Menampilkan detail Kop Surat"""
    kop_surat = get_object_or_404(KopSurat, pk=pk)
    return render(request, 'asn_app/kop_surat_detail.html', {'kop_surat': kop_surat})

def kop_surat_create(request):
    """Membuat Kop Surat baru"""
    if request.method == 'POST':
        form = KopSuratForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('kop_surat_list')
    else:
        form = KopSuratForm()
    return render(request, 'asn_app/kop_surat_form.html', {'form': form, 'title': 'Tambah Kop Surat'})

def kop_surat_update(request, pk):
    """Mengupdate data Kop Surat"""
    kop_surat = get_object_or_404(KopSurat, pk=pk)
    if request.method == 'POST':
        form = KopSuratForm(request.POST, request.FILES, instance=kop_surat)
        if form.is_valid():
            form.save()
            return redirect('kop_surat_detail', pk=kop_surat.pk)
    else:
        form = KopSuratForm(instance=kop_surat)
    return render(request, 'asn_app/kop_surat_form.html', {'form': form, 'title': 'Edit Kop Surat'})

def kop_surat_delete(request, pk):
    """Menghapus data Kop Surat"""
    kop_surat = get_object_or_404(KopSurat, pk=pk)
    if request.method == 'POST':
        kop_surat.delete()
        return redirect('kop_surat_list')
    return render(request, 'asn_app/kop_surat_confirm_delete.html', {'kop_surat': kop_surat})

def export_asn_excel(request):
    """Export all ASN data to an Excel file."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    asns = ASN.objects.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'ASN Data'

    # Write headers
    headers = [
        'NIP', 'Nama', 'Tempat Lahir', 'Tanggal Lahir', 'Jenis Kelamin',
        'Agama', 'Alamat', 'Email', 'Telepon', 'Jabatan', 'Pangkat',
        'Golongan', 'Unit Kerja'
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Write data
    for row_num, asn in enumerate(asns, 2):
        sheet[f'A{row_num}'] = asn.nip
        sheet[f'B{row_num}'] = asn.nama
        sheet[f'C{row_num}'] = asn.tempat_lahir
        sheet[f'D{row_num}'] = asn.tanggal_lahir.strftime('%Y-%m-%d') if asn.tanggal_lahir else ''
        sheet[f'E{row_num}'] = asn.get_jenis_kelamin_display()
        sheet[f'F{row_num}'] = asn.agama
        sheet[f'G{row_num}'] = asn.alamat
        sheet[f'H{row_num}'] = asn.email
        sheet[f'I{row_num}'] = asn.telepon
        sheet[f'J{row_num}'] = asn.jabatan
        sheet[f'K{row_num}'] = asn.pangkat
        sheet[f'L{row_num}'] = asn.golongan
        sheet[f'M{row_num}'] = asn.unit_kerja

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=asn_data.xlsx'
    workbook.save(response)
    return response

def import_asn_excel(request):
    """Import ASN data from an Excel file."""
    if request.method == 'POST':
        import openpyxl
        from datetime import datetime

        excel_file = request.FILES['excel_file']
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            (
                nip, nama, tempat_lahir, tanggal_lahir_val, jenis_kelamin,
                agama, alamat, email, telepon, jabatan, pangkat,
                golongan, unit_kerja
            ) = row

            # Handle date conversion
            tanggal_lahir = None
            if isinstance(tanggal_lahir_val, datetime):
                tanggal_lahir = tanggal_lahir_val.date()
            elif isinstance(tanggal_lahir_val, str):
                try:
                    tanggal_lahir = datetime.strptime(tanggal_lahir_val, '%Y-%m-%d').date()
                except ValueError:
                    # Handle other date formats if necessary
                    pass

            # Handle choices
            jenis_kelamin_code = jenis_kelamin[0].upper() if jenis_kelamin else ''
            agama_code = agama.upper() if agama else ''

            ASN.objects.create(
                nip=nip,
                nama=nama,
                tempat_lahir=tempat_lahir,
                tanggal_lahir=tanggal_lahir,
                jenis_kelamin=jenis_kelamin_code,
                agama=agama_code,
                alamat=alamat,
                email=email,
                telepon=telepon,
                jabatan=jabatan,
                pangkat=pangkat,
                golongan=golongan,
                unit_kerja=unit_kerja
            )
        return redirect('asn_list')
    return render(request, 'asn_app/import_form.html')


# Surat Santunan Korpri Views
def surat_santunan_korpri_list(request):
    surat_list = SuratSantunanKorpri.objects.all().order_by('-created_at')
    return render(request, 'asn_app/surat_santunan_korpri_list.html', {'surat_list': surat_list})

def surat_santunan_korpri_detail(request, pk):
    surat = get_object_or_404(SuratSantunanKorpri, pk=pk)
    return render(request, 'asn_app/surat_santunan_korpri_detail.html', {'surat': surat})

def surat_santunan_korpri_create(request):
    if request.method == 'POST':
        form = SuratSantunanKorpriForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('surat_santunan_korpri_list')
    else:
        form = SuratSantunanKorpriForm()
    return render(request, 'asn_app/surat_santunan_korpri_form.html', {'form': form, 'title': 'Tambah Surat Santunan Korpri'})

def surat_santunan_korpri_update(request, pk):
    surat = get_object_or_404(SuratSantunanKorpri, pk=pk)
    if request.method == 'POST':
        form = SuratSantunanKorpriForm(request.POST, instance=surat)
        if form.is_valid():
            form.save()
            return redirect('surat_santunan_korpri_detail', pk=surat.pk)
    else:
        form = SuratSantunanKorpriForm(instance=surat)
    return render(request, 'asn_app/surat_santunan_korpri_form.html', {'form': form, 'title': 'Edit Surat Santunan Korpri'})

def surat_santunan_korpri_delete(request, pk):
    surat = get_object_or_404(SuratSantunanKorpri, pk=pk)
    if request.method == 'POST':
        surat.delete()
        return redirect('surat_santunan_korpri_list')
    return render(request, 'asn_app/surat_santunan_korpri_confirm_delete.html', {'surat': surat})

def surat_santunan_korpri_export_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    surat = get_object_or_404(SuratSantunanKorpri, pk=pk)

    kop_surat_base64 = None
    if surat.kop_surat and surat.kop_surat.gambar:
        if os.path.exists(surat.kop_surat.gambar.path):
            try:
                with open(surat.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64 for large PDF: {e}")

    html_string = render_to_string('asn_app/surat_santunan_korpri_pdf_template.html', {
        'surat': surat,
        'request': request,
        'kop_surat_base64': kop_surat_base64,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=surat_santunan_{surat.nomor_surat}.pdf'

    return response

# Nota Dinas Views
def nota_dinas_list(request):
    nota_dinas_list = NotaDinas.objects.all().order_by('-tanggal')
    return render(request, 'asn_app/nota_dinas_list.html', {'nota_dinas_list': nota_dinas_list})

def nota_dinas_detail(request, pk):
    nota_dinas = get_object_or_404(NotaDinas, pk=pk)
    return render(request, 'asn_app/nota_dinas_detail.html', {'nota_dinas': nota_dinas})

def nota_dinas_create(request):
    if request.method == 'POST':
        form = NotaDinasForm(request.POST)
        formset = PesertaNotaDinasFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            nota_dinas = form.save()
            formset.instance = nota_dinas
            formset.save()
            return redirect('nota_dinas_list')
    else:
        form = NotaDinasForm()
        formset = PesertaNotaDinasFormSet()
    
    kelas_list = Siswa.objects.exclude(kelas__isnull=True).exclude(kelas__exact='').values_list('kelas', flat=True).distinct().order_by('kelas')
    jurusan_list = Siswa.objects.exclude(jurusan__isnull=True).exclude(jurusan__exact='').values_list('jurusan', flat=True).distinct().order_by('jurusan')
    
    return render(request, 'asn_app/nota_dinas_form.html', {
        'form': form, 'formset': formset, 'title': 'Tambah Nota Dinas',
        'kelas_list': kelas_list, 'jurusan_list': jurusan_list
    })

def nota_dinas_update(request, pk):
    nota_dinas = get_object_or_404(NotaDinas, pk=pk)
    if request.method == 'POST':
        form = NotaDinasForm(request.POST, instance=nota_dinas)
        formset = PesertaNotaDinasFormSet(request.POST, instance=nota_dinas)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('nota_dinas_detail', pk=nota_dinas.pk)
    else:
        form = NotaDinasForm(instance=nota_dinas)
        formset = PesertaNotaDinasFormSet(instance=nota_dinas)
    
    kelas_list = Siswa.objects.exclude(kelas__isnull=True).exclude(kelas__exact='').values_list('kelas', flat=True).distinct().order_by('kelas')
    jurusan_list = Siswa.objects.exclude(jurusan__isnull=True).exclude(jurusan__exact='').values_list('jurusan', flat=True).distinct().order_by('jurusan')
    
    return render(request, 'asn_app/nota_dinas_form.html', {
        'form': form, 'formset': formset, 'title': 'Edit Nota Dinas',
        'kelas_list': kelas_list, 'jurusan_list': jurusan_list
    })

def nota_dinas_delete(request, pk):
    nota_dinas = get_object_or_404(NotaDinas, pk=pk)
    if request.method == 'POST':
        nota_dinas.delete()
        return redirect('nota_dinas_list')
    return render(request, 'asn_app/nota_dinas_confirm_delete.html', {'nota_dinas': nota_dinas})

def peserta_nota_dinas_list(request):
    nota_dinas_id = request.GET.get('nota_dinas')
    peserta_list = PesertaNotaDinas.objects.select_related('nota_dinas', 'pegawai', 'siswa').all().order_by('nota_dinas__tanggal', 'id')
    nota_dinas = None
    if nota_dinas_id:
        peserta_list = peserta_list.filter(nota_dinas_id=nota_dinas_id)
        nota_dinas = get_object_or_404(NotaDinas, pk=nota_dinas_id)
    return render(request, 'asn_app/peserta_nota_dinas_list.html', {
        'peserta_list': peserta_list,
        'nota_dinas': nota_dinas,
    })

def peserta_nota_dinas_create(request):
    next_url = request.GET.get('next') or request.POST.get('next') or reverse('peserta_nota_dinas_list')
    if request.method == 'POST':
        form = PesertaNotaDinasCRUDForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(next_url)
    else:
        form = PesertaNotaDinasCRUDForm()
        nota_dinas_id = request.GET.get('nota_dinas')
        if nota_dinas_id:
            form.fields['nota_dinas'].initial = nota_dinas_id
    return render(request, 'asn_app/peserta_nota_dinas_form.html', {
        'form': form, 'title': 'Tambah Peserta dan Pendamping Kegiatan LKS', 'next_url': next_url
    })

def peserta_nota_dinas_update(request, pk):
    peserta = get_object_or_404(PesertaNotaDinas.objects.select_related('nota_dinas', 'pegawai', 'siswa'), pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or reverse('peserta_nota_dinas_list')
    if request.method == 'POST':
        form = PesertaNotaDinasCRUDForm(request.POST, instance=peserta)
        if form.is_valid():
            form.save()
            return redirect(next_url)
    else:
        form = PesertaNotaDinasCRUDForm(instance=peserta)
    return render(request, 'asn_app/peserta_nota_dinas_form.html', {
        'form': form, 'title': 'Edit Peserta dan Pendamping Kegiatan LKS', 'next_url': next_url, 'peserta': peserta
    })

def peserta_nota_dinas_delete(request, pk):
    peserta = get_object_or_404(PesertaNotaDinas.objects.select_related('nota_dinas', 'pegawai', 'siswa'), pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or reverse('peserta_nota_dinas_list')
    if request.method == 'POST':
        peserta.delete()
        return redirect(next_url)
    return render(request, 'asn_app/peserta_nota_dinas_confirm_delete.html', {
        'peserta': peserta, 'next_url': next_url
    })


def nota_dinas_export_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    nota_dinas = get_object_or_404(NotaDinas, pk=pk)

    kop_surat_base64 = None
    if nota_dinas.kop_surat and nota_dinas.kop_surat.gambar:
        if os.path.exists(nota_dinas.kop_surat.gambar.path):
            try:
                with open(nota_dinas.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = nota_dinas.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64 for large PDF: {e}")

    jumlah_peserta = nota_dinas.peserta_nota_dinas.count()
    if jumlah_peserta == 0:
        jumlah_peserta = nota_dinas.pegawai.count() + nota_dinas.siswa.count()
    show_lampiran_link = jumlah_peserta > 3
    lampiran_url = request.build_absolute_uri(reverse('nota_dinas_lampiran_pdf', kwargs={'pk': nota_dinas.pk}))

    html_string = render_to_string('asn_app/nota_dinas_pdf_template.html', {
        'nota_dinas': nota_dinas,
        'request': request,
        'kop_surat_base64': kop_surat_base64,
        'show_lampiran_link': show_lampiran_link,
        'lampiran_url': lampiran_url,
        'jumlah_peserta': jumlah_peserta,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    return response


def nota_dinas_lampiran_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    nota_dinas = get_object_or_404(NotaDinas, pk=pk)

    kop_surat_base64 = None
    if nota_dinas.kop_surat and nota_dinas.kop_surat.gambar:
        if os.path.exists(nota_dinas.kop_surat.gambar.path):
            try:
                with open(nota_dinas.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = nota_dinas.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64 for lampiran PDF: {e}")

    html_string = render_to_string('asn_app/nota_dinas_lampiran_template.html', {
        'nota_dinas': nota_dinas,
        'request': request,
        'kop_surat_base64': kop_surat_base64,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing lampiran PDF: {e}")
        return HttpResponse(f"Error writing lampiran PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="lampiran_nota_dinas_{nota_dinas.pk}.pdf"'
    return response


# Surat Usulan Views
def surat_usulan_list(request):
    surat_usulan_list = SuratUsulan.objects.all().order_by('-tanggal')
    return render(request, 'asn_app/surat_usulan_list.html', {'surat_usulan_list': surat_usulan_list})

def surat_usulan_detail(request, pk):
    surat_usulan = get_object_or_404(SuratUsulan, pk=pk)
    return render(request, 'asn_app/surat_usulan_detail.html', {'surat_usulan': surat_usulan})

def surat_usulan_create(request):
    if request.method == 'POST':
        form = SuratUsulanForm(request.POST)
        formset = PesertaSuratUsulanFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            surat_usulan = form.save()
            formset.instance = surat_usulan
            formset.save()
            return redirect('surat_usulan_list')
    else:
        form = SuratUsulanForm()
        formset = PesertaSuratUsulanFormSet()
    
    return render(request, 'asn_app/surat_usulan_form.html', {
        'form': form, 'formset': formset, 'title': 'Tambah Surat Usulan'
    })

def surat_usulan_update(request, pk):
    surat_usulan = get_object_or_404(SuratUsulan, pk=pk)
    if request.method == 'POST':
        form = SuratUsulanForm(request.POST, instance=surat_usulan)
        formset = PesertaSuratUsulanFormSet(request.POST, instance=surat_usulan)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('surat_usulan_detail', pk=surat_usulan.pk)
    else:
        form = SuratUsulanForm(instance=surat_usulan)
        formset = PesertaSuratUsulanFormSet(instance=surat_usulan)
    
    return render(request, 'asn_app/surat_usulan_form.html', {
        'form': form, 'formset': formset, 'title': 'Edit Surat Usulan'
    })

def surat_usulan_delete(request, pk):
    surat_usulan = get_object_or_404(SuratUsulan, pk=pk)
    if request.method == 'POST':
        surat_usulan.delete()
        return redirect('surat_usulan_list')
    return render(request, 'asn_app/surat_usulan_confirm_delete.html', {'surat_usulan': surat_usulan})

def peserta_surat_usulan_list(request):
    surat_usulan_id = request.GET.get('surat_usulan')
    peserta_list = PesertaSuratUsulan.objects.select_related('surat_usulan', 'pegawai').all().order_by('surat_usulan__tanggal', 'id')
    surat_usulan = None
    if surat_usulan_id:
        peserta_list = peserta_list.filter(surat_usulan_id=surat_usulan_id)
        surat_usulan = get_object_or_404(SuratUsulan, pk=surat_usulan_id)
    return render(request, 'asn_app/peserta_surat_usulan_list.html', {
        'peserta_list': peserta_list,
        'surat_usulan': surat_usulan,
    })

def peserta_surat_usulan_create(request):
    next_url = request.GET.get('next') or request.POST.get('next') or reverse('peserta_surat_usulan_list')
    if request.method == 'POST':
        form = PesertaSuratUsulanCRUDForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(next_url)
    else:
        form = PesertaSuratUsulanCRUDForm()
        surat_usulan_id = request.GET.get('surat_usulan')
        if surat_usulan_id:
            form.fields['surat_usulan'].initial = surat_usulan_id
    return render(request, 'asn_app/peserta_surat_usulan_form.html', {
        'form': form, 'title': 'Tambah Peserta Surat Usulan', 'next_url': next_url
    })

def peserta_surat_usulan_update(request, pk):
    peserta = get_object_or_404(PesertaSuratUsulan.objects.select_related('surat_usulan', 'pegawai'), pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or reverse('peserta_surat_usulan_list')
    if request.method == 'POST':
        form = PesertaSuratUsulanCRUDForm(request.POST, instance=peserta)
        if form.is_valid():
            form.save()
            return redirect(next_url)
    else:
        form = PesertaSuratUsulanCRUDForm(instance=peserta)
    return render(request, 'asn_app/peserta_surat_usulan_form.html', {
        'form': form, 'title': 'Edit Peserta Surat Usulan', 'next_url': next_url, 'peserta': peserta
    })

def peserta_surat_usulan_delete(request, pk):
    peserta = get_object_or_404(PesertaSuratUsulan.objects.select_related('surat_usulan', 'pegawai'), pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or reverse('peserta_surat_usulan_list')
    if request.method == 'POST':
        peserta.delete()
        return redirect(next_url)
    return render(request, 'asn_app/peserta_surat_usulan_confirm_delete.html', {
        'peserta': peserta, 'next_url': next_url
    })

def surat_usulan_export_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    surat_usulan = get_object_or_404(SuratUsulan, pk=pk)

    kop_surat_base64 = None
    if surat_usulan.kop_surat and surat_usulan.kop_surat.gambar:
        if os.path.exists(surat_usulan.kop_surat.gambar.path):
            try:
                with open(surat_usulan.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat_usulan.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64 for Surat Usulan PDF: {e}")

    jumlah_peserta = surat_usulan.peserta_surat_usulan.count()
    show_lampiran_link = jumlah_peserta > 0
    lampiran_url = request.build_absolute_uri(reverse('surat_usulan_lampiran_pdf', kwargs={'pk': surat_usulan.pk}))

    html_string = render_to_string('asn_app/surat_usulan_pdf_template.html', {
        'surat_usulan': surat_usulan,
        'request': request,
        'kop_surat_base64': kop_surat_base64,
        'show_lampiran_link': show_lampiran_link,
        'lampiran_url': lampiran_url,
        'jumlah_peserta': jumlah_peserta,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF for Surat Usulan: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=surat_usulan_{surat_usulan.pk}.pdf'
    return response


def surat_usulan_lampiran_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    surat_usulan = get_object_or_404(SuratUsulan, pk=pk)

    kop_surat_base64 = None
    if surat_usulan.kop_surat and surat_usulan.kop_surat.gambar:
        if os.path.exists(surat_usulan.kop_surat.gambar.path):
            try:
                with open(surat_usulan.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat_usulan.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64 for Surat Usulan Lampiran PDF: {e}")

    html_string = render_to_string('asn_app/surat_usulan_lampiran_template.html', {
        'surat_usulan': surat_usulan,
        'request': request,
        'kop_surat_base64': kop_surat_base64,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing lampiran PDF for Surat Usulan: {e}")
        return HttpResponse(f"Error writing lampiran PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=lampiran_surat_usulan_{surat_usulan.pk}.pdf'
    return response


# ST Satyalancana Views
def st_satyalancana_list(request):
    st_list = StSatyalancana.objects.all().order_by('-tanggal')
    return render(request, 'asn_app/st_satyalancana_list.html', {'st_list': st_list})

def st_satyalancana_detail(request, pk):
    st = get_object_or_404(StSatyalancana, pk=pk)
    return render(request, 'asn_app/st_satyalancana_detail.html', {'st': st})

def st_satyalancana_create(request):
    if request.method == 'POST':
        form = StSatyalancanaForm(request.POST)
        if form.is_valid():
            st = form.save(commit=False)
            if st.kepada_pegawai and not st.kepada_nama:
                st.kepada_nama = st.kepada_pegawai.nama
            if st.kepada_pegawai and not st.kepada_nip:
                st.kepada_nip = st.kepada_pegawai.nip
            if st.kepada_pegawai and not st.kepada_pangkat_gol:
                st.kepada_pangkat_gol = f"{st.kepada_pegawai.pangkat} / {st.kepada_pegawai.golongan}"
            if st.kepada_pegawai and not st.kepada_jabatan:
                st.kepada_jabatan = st.kepada_pegawai.jabatan
            if st.kepada_pegawai and not st.kepada_unit_kerja:
                st.kepada_unit_kerja = st.kepada_pegawai.unit_kerja
            st.save()
            return redirect('st_satyalancana_list')
    else:
        form = StSatyalancanaForm()
    return render(request, 'asn_app/st_satyalancana_form.html', {
        'form': form, 'title': 'Tambah Surat Tugas Satyalancana'
    })

def st_satyalancana_update(request, pk):
    st = get_object_or_404(StSatyalancana, pk=pk)
    if request.method == 'POST':
        form = StSatyalancanaForm(request.POST, instance=st)
        if form.is_valid():
            st = form.save(commit=False)
            if st.kepada_pegawai and not st.kepada_nama:
                st.kepada_nama = st.kepada_pegawai.nama
            if st.kepada_pegawai and not st.kepada_nip:
                st.kepada_nip = st.kepada_pegawai.nip
            if st.kepada_pegawai and not st.kepada_pangkat_gol:
                st.kepada_pangkat_gol = f"{st.kepada_pegawai.pangkat} / {st.kepada_pegawai.golongan}"
            if st.kepada_pegawai and not st.kepada_jabatan:
                st.kepada_jabatan = st.kepada_pegawai.jabatan
            if st.kepada_pegawai and not st.kepada_unit_kerja:
                st.kepada_unit_kerja = st.kepada_pegawai.unit_kerja
            st.save()
            return redirect('st_satyalancana_detail', pk=st.pk)
    else:
        form = StSatyalancanaForm(instance=st)
    return render(request, 'asn_app/st_satyalancana_form.html', {
        'form': form, 'title': 'Edit Surat Tugas Satyalancana'
    })

def st_satyalancana_delete(request, pk):
    st = get_object_or_404(StSatyalancana, pk=pk)
    if request.method == 'POST':
        st.delete()
        return redirect('st_satyalancana_list')
    return render(request, 'asn_app/st_satyalancana_confirm_delete.html', {'st': st})

def st_satyalancana_export_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    st = get_object_or_404(StSatyalancana, pk=pk)

    kop_surat_base64 = None
    if st.kop_surat and st.kop_surat.gambar:
        if os.path.exists(st.kop_surat.gambar.path):
            try:
                with open(st.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = st.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64 for ST Satyalancana PDF: {e}")

    html_string = render_to_string('asn_app/st_satyalancana_pdf_template.html', {
        'st': st,
        'request': request,
        'kop_surat_base64': kop_surat_base64,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF for ST Satyalancana: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=st_satyalancana_{st.pk}.pdf'
    return response


# DRH Satyalancana Views
def drh_satyalancana_list(request):
    drh_list = DRHSatyalancana.objects.select_related('asn').all().order_by('-created_at')
    return render(request, 'asn_app/drh_satyalancana_list.html', {'drh_list': drh_list})

def drh_satyalancana_detail(request, pk):
    drh = get_object_or_404(DRHSatyalancana.objects.select_related('asn'), pk=pk)
    return render(request, 'asn_app/drh_satyalancana_detail.html', {'drh': drh})

def drh_satyalancana_create(request):
    if request.method == 'POST':
        form = DRHSatyalancanaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('drh_satyalancana_list')
    else:
        form = DRHSatyalancanaForm()
    return render(request, 'asn_app/drh_satyalancana_form.html', {
        'form': form, 'title': 'Tambah DRH Satyalancana'
    })

def drh_satyalancana_update(request, pk):
    drh = get_object_or_404(DRHSatyalancana.objects.select_related('asn'), pk=pk)
    if request.method == 'POST':
        form = DRHSatyalancanaForm(request.POST, instance=drh)
        if form.is_valid():
            form.save()
            return redirect('drh_satyalancana_detail', pk=drh.pk)
    else:
        form = DRHSatyalancanaForm(instance=drh)
    return render(request, 'asn_app/drh_satyalancana_form.html', {
        'form': form, 'title': 'Edit DRH Satyalancana'
    })

def drh_satyalancana_delete(request, pk):
    drh = get_object_or_404(DRHSatyalancana.objects.select_related('asn'), pk=pk)
    if request.method == 'POST':
        drh.delete()
        return redirect('drh_satyalancana_list')
    return render(request, 'asn_app/drh_satyalancana_confirm_delete.html', {'drh': drh})

def drh_satyalancana_export_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    drh = get_object_or_404(DRHSatyalancana.objects.select_related('asn'), pk=pk)
    asn = drh.asn

    kop_surat_base64 = None
    if asn.unit_kerja:
        # Try to get kop surat from StSatyalancana or use a default
        kop_surat = StSatyalancana.objects.first().kop_surat if StSatyalancana.objects.exists() else None
        if kop_surat and kop_surat.gambar:
            if os.path.exists(kop_surat.gambar.path):
                try:
                    with open(kop_surat.gambar.path, 'rb') as image_file:
                        image_data = image_file.read()
                        image_format = kop_surat.gambar.name.split('.')[-1].lower()
                        if image_format == 'jpg':
                            image_format = 'jpeg'
                        kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
                except Exception as e:
                    logging.error(f"Error encoding image to base64 for DRH PDF: {e}")

    html_string = render_to_string('asn_app/drh_satyalancana_pdf_template.html', {
        'drh': drh,
        'asn': asn,
        'request': request,
        'kop_surat_base64': kop_surat_base64,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF for DRH Satyalancana: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=drh_satyalancana_{drh.pk}.pdf'
    return response


# Hari Libur Views
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy

class HariLiburListView(ListView):
    model = HariLibur
    template_name = 'asn_app/hari_libur_list.html'
    context_object_name = 'hari_libur_list'
    paginate_by = 10

class HariLiburDetailView(DetailView):
    model = HariLibur
    template_name = 'asn_app/hari_libur_detail.html'
    context_object_name = 'hari_libur'

class HariLiburCreateView(CreateView):
    model = HariLibur
    form_class = HariLiburForm
    template_name = 'asn_app/hari_libur_form.html'
    success_url = reverse_lazy('hari_libur_list')

class HariLiburUpdateView(UpdateView):
    model = HariLibur
    form_class = HariLiburForm
    template_name = 'asn_app/hari_libur_form.html'
    context_object_name = 'hari_libur'

    def get_success_url(self):
        return reverse_lazy('hari_libur_detail', kwargs={'pk': self.object.pk})

class HariLiburDeleteView(DeleteView):
    model = HariLibur
    template_name = 'asn_app/hari_libur_confirm_delete.html'
    success_url = reverse_lazy('hari_libur_list')
    context_object_name = 'hari_libur'

# Surat Cuti Views
class SuratCutiListView(ListView):
    model = SuratCuti
    template_name = 'asn_app/surat_cuti_list.html'
    context_object_name = 'surat_cuti_list'
    paginate_by = 7

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('nama', '')
        if search_query:
            queryset = queryset.filter(pegawai__nama__icontains=search_query)
        return queryset

class SuratCutiDetailView(DetailView):
    model = SuratCuti
    template_name = 'asn_app/surat_cuti_detail.html'
    context_object_name = 'surat_cuti'

class SuratCutiCreateView(CreateView):
    model = SuratCuti
    form_class = SuratCutiForm
    template_name = 'asn_app/surat_cuti_form.html'
    success_url = reverse_lazy('surat_cuti_list')

    def get_initial(self):
        initial = super().get_initial()
        pegawai_id = self.request.GET.get('pegawai')
        if pegawai_id:
            initial['pegawai'] = pegawai_id
        initial['alasan_cuti'] = 'Keperluan keluarga'
        initial['perihal_surat'] = 'Permohonan'
        initial['tujuan_surat'] = 'Kepala SMK Negeri 1 Koba'
        initial['tempat_ditetapkan'] = 'Koba'
        penandatangan = ASN.objects.filter(nama__icontains="SYAHRYANTO,S.T.,M.Pd").first()
        if penandatangan:
            initial['penandatangan'] = penandatangan.pk
        kop_surat = KopSurat.objects.filter(gambar__icontains="kop_smk1_koba.png").first()
        if kop_surat:
            initial['kop_surat'] = kop_surat.pk
        return initial

class SuratCutiUpdateView(UpdateView):
    model = SuratCuti
    form_class = SuratCutiForm
    template_name = 'asn_app/surat_cuti_form.html'
    context_object_name = 'surat_cuti'

    def get_success_url(self):
        return reverse_lazy('surat_cuti_detail', kwargs={'pk': self.object.pk})

class SuratCutiDeleteView(DeleteView):
    model = SuratCuti
    template_name = 'asn_app/surat_cuti_confirm_delete.html'
    success_url = reverse_lazy('surat_cuti_list')
    context_object_name = 'surat_cuti'

# Sisa Cuti Views
class SisaCutiListView(ListView):
    model = SisaCuti
    template_name = 'asn_app/sisa_cuti_list.html'
    context_object_name = 'sisa_cuti_list'
    paginate_by = 7

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(pegawai__nama__icontains=search_query)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context

class SisaCutiDetailView(DetailView):
    model = SisaCuti
    template_name = 'asn_app/sisa_cuti_detail.html'
    context_object_name = 'sisa_cuti'

class SisaCutiCreateView(CreateView):
    model = SisaCuti
    form_class = SisaCutiForm
    template_name = 'asn_app/sisa_cuti_form.html'
    success_url = reverse_lazy('sisa_cuti_list')

    def get_initial(self):
        """Pre-fill form with default initial allocation values."""
        initial = super().get_initial()
        pegawai_id = self.request.GET.get('pegawai_id')

        # Set default initial allocation values
        initial['initial_tahun_n'] = 12
        initial['initial_tahun_n_1'] = 6
        initial['initial_tahun_n_2'] = 6

        if pegawai_id:
            from .models import ASN
            try:
                asn = ASN.objects.get(pk=pegawai_id)
                # Check if SisaCuti already exists
                sisa_cuti = SisaCuti.objects.filter(pegawai=asn).first()
                if sisa_cuti:
                    # Use existing values if SisaCuti exists
                    initial['initial_tahun_n'] = sisa_cuti.initial_tahun_n
                    initial['initial_tahun_n_1'] = sisa_cuti.initial_tahun_n_1
                    initial['initial_tahun_n_2'] = sisa_cuti.initial_tahun_n_2
            except ASN.DoesNotExist:
                pass

        return initial

    def form_valid(self, form):
        """Save initial allocation and calculate sisa cuti."""
        instance = form.save(commit=False)
        # Recalculate sisa cuti based on initial allocation values
        instance.calculate_sisa_cuti_from_surat()
        instance.save()
        return super().form_valid(form)

class SisaCutiUpdateView(UpdateView):
    model = SisaCuti
    form_class = SisaCutiForm
    template_name = 'asn_app/sisa_cuti_form.html'
    context_object_name = 'sisa_cuti'

    def get_success_url(self):
        return reverse_lazy('sisa_cuti_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        """Save initial allocation and recalculate sisa cuti."""
        # Save the form first
        instance = form.save(commit=False)

        # Recalculate sisa cuti based on new initial allocation values
        instance.calculate_sisa_cuti_from_surat()
        instance.save()

        return super().form_valid(form)

class SisaCutiDeleteView(DeleteView):
    model = SisaCuti
    template_name = 'asn_app/sisa_cuti_confirm_delete.html'
    success_url = reverse_lazy('sisa_cuti_list')
    context_object_name = 'sisa_cuti'

def surat_cuti_export_pdf(request, pk):
    import os
    import base64
    from datetime import date, timedelta
    from weasyprint import HTML

    surat_cuti = get_object_or_404(SuratCuti, pk=pk)

    logger = logging.getLogger(__name__)
    logger.info(f"Generating PDF for SuratCuti: {surat_cuti}")

    # Calculate leave duration
    try:
        leave_days = surat_cuti.calculate_effective_leave_days()

        # Define Indonesian month names
        month_names = {
            'January': 'Januari', 'February': 'Februari', 'March': 'Maret', 'April': 'April',
            'May': 'Mei', 'June': 'Juni', 'July': 'Juli', 'August': 'Agustus',
            'September': 'September', 'October': 'Oktober', 'November': 'November', 'December': 'Desember'
        }

        # Format dates with English month names first, then replace with Indonesian names
        start_date_str = surat_cuti.tanggal_awal.strftime('%d %B %Y')
        end_date_str = surat_cuti.tanggal_akhir.strftime('%d %B %Y')

        # Replace English month names with Indonesian
        for eng_month, indo_month in month_names.items():
            start_date_str = start_date_str.replace(eng_month, indo_month)
            end_date_str = end_date_str.replace(eng_month, indo_month)

        # Convert number to Indonesian words
        number_words = {
            1: "satu", 2: "dua", 3: "tiga", 4: "empat", 5: "lima",
            6: "enam", 7: "tujuh", 8: "delapan", 9: "sembilan", 10: "sepuluh",
            11: "sebelas", 12: "dua belas", 13: "tiga belas", 14: "empat belas",
            15: "lima belas", 16: "enam belas", 17: "tujuh belas", 18: "delapan belas",
            19: "sembilan belas", 20: "dua puluh"
        }

        # For leave > 30 days, show as "3 (tiga) bulan"
        if leave_days > 30:
            written_number = "tiga"
            leave_duration_text = f"selama 3 ({written_number}) bulan, terhitung mulai tanggal {start_date_str} sampai dengan tanggal {end_date_str}"
        elif surat_cuti.tanggal_awal == surat_cuti.tanggal_akhir:
            if leave_days in number_words:
                written_number = number_words[leave_days]
            else:
                written_number = str(leave_days)
            leave_duration_text = f"selama {leave_days} ({written_number}) hari kerja pada tanggal {start_date_str}"
        else:
            if leave_days in number_words:
                written_number = number_words[leave_days]
            else:
                written_number = str(leave_days)
            leave_duration_text = f"selama {leave_days} ({written_number}) hari kerja, terhitung mulai tanggal {start_date_str} sampai dengan tanggal {end_date_str}"
        logger.info(f"Leave duration: {leave_duration_text}")
    except Exception as e:
        logger.error(f"Error calculating leave duration: {e}")
        leave_duration_text = "durasi cuti tidak dapat dihitung"

    # Handle kop surat
    kop_surat_base64 = None
    if surat_cuti.kop_surat and surat_cuti.kop_surat.gambar:
        logger.info(f"Kop surat: {surat_cuti.kop_surat}, gambar: {surat_cuti.kop_surat.gambar}")
        if os.path.exists(surat_cuti.kop_surat.gambar.path):
            try:
                with open(surat_cuti.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat_cuti.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
                    logger.info("Kop surat encoded successfully")
            except Exception as e:
                logger.error(f"Error encoding image to base64: {e}")
        else:
            logger.error("Kop surat file does not exist")

    try:
        html_string = render_to_string('asn_app/surat_cuti_pdf_template.html', {
            'surat_cuti': surat_cuti,
            'leave_duration_text': leave_duration_text,
            'kop_surat_base64': kop_surat_base64,
            'request': request,
        })
        logger.info("HTML template rendered successfully")
        # Write rendered HTML to a debug file to inspect template rendering
        try:
            from django.conf import settings
            debug_path = getattr(settings, 'BASE_DIR', '.')
            debug_file = os.path.join(debug_path, f'debug_surat_cuti_{surat_cuti.pk}.html')
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write(html_string)
            logger.info(f"Wrote debug HTML to {debug_file}")
        except Exception as e:
            logger.error(f"Error writing debug HTML file: {e}")
    except Exception as e:
        logger.error(f"Error rendering template: {e}")
        return HttpResponse(f"Error rendering template: {e}", status=500)

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
        logger.info(f"PDF generated successfully, size: {len(result)} bytes")
    except Exception as e:
        logger.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    # Sanitize filename to avoid issues with special characters
    safe_name = "".join(c for c in surat_cuti.pegawai.nama if c.isalnum() or c in (' ', '-', '_')).rstrip()
    filename = f'surat_cuti_{safe_name}.pdf'

    response = HttpResponse(result, content_type='application/pdf')
    # Explicitly delete any existing Content-Disposition header before setting our own
    if 'Content-Disposition' in response:
        del response['Content-Disposition']
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

def laporan_cuti_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    asn = get_object_or_404(ASN, pk=pk)

    # Get year parameters from request (supports multiple years)
    years = request.GET.getlist('years')

    # If no years selected, default to current year
    if not years:
        years = [str(now().year)]

    # Convert to integers
    try:
        years = [int(y) for y in years]
    except (ValueError, TypeError):
        years = [now().year]

    # Get all surat_cuti for selected years
    surat_cuti_queryset = SuratCuti.objects.filter(
        pegawai=asn,
        tanggal_awal__year__in=years
    ).order_by('tanggal_awal')

    sisa_cuti_obj = SisaCuti.objects.filter(pegawai=asn).first()

    # Initialize running balances with INITIAL allocation values (not remaining)
    initial_sisa_tahun_n = sisa_cuti_obj.initial_tahun_n if sisa_cuti_obj else 0
    initial_sisa_tahun_n_1 = sisa_cuti_obj.initial_tahun_n_1 if sisa_cuti_obj else 0
    initial_sisa_tahun_n_2 = sisa_cuti_obj.initial_tahun_n_2 if sisa_cuti_obj else 0
    initial_total_sisa_cuti = (initial_sisa_tahun_n + initial_sisa_tahun_n_1 + initial_sisa_tahun_n_2) if sisa_cuti_obj else 0

    # Create a list to hold processed leave data
    processed_surat_cuti_list = []

    # Start with the initial balances for the first row's "before leave" state
    # First row ATB should show the full initial balance (not decreased)
    current_sisa_tahun_n_balance = initial_sisa_tahun_n
    current_total_sisa_cuti_balance = initial_total_sisa_cuti

    for idx, surat_cuti in enumerate(surat_cuti_queryset):
        lhc = surat_cuti.calculate_effective_leave_days()

        # ATB for this row:
        # - First row: show the full initial balance (not decreased)
        # - Subsequent rows: show the balance after previous leaves
        if idx == 0:
            # First row - ATB is the initial total (undecreased)
            atb_for_row = initial_total_sisa_cuti
        else:
            # Subsequent rows - ATB is the running balance
            atb_for_row = current_total_sisa_cuti_balance

        # STB for this row is ATB for this row minus LHC for this row
        stb_for_row = atb_for_row - lhc

        # Update running balances for the *next* iteration
        current_sisa_tahun_n_balance -= lhc
        current_total_sisa_cuti_balance -= lhc

        processed_surat_cuti_list.append({
            'surat_cuti': surat_cuti,
            'lhc': lhc,
            'atb_for_row': atb_for_row,
            'stb_for_row': stb_for_row,
        })

    rows_per_page = 22
    pages = []
    for start_index in range(0, len(processed_surat_cuti_list), rows_per_page):
        page_rows = []
        for row_index, item in enumerate(
            processed_surat_cuti_list[start_index:start_index + rows_per_page],
            start=start_index + 1,
        ):
            row = item.copy()
            row['number'] = row_index
            page_rows.append(row)
        pages.append({'rows': page_rows})

    # Default penandatangan: cari pegawai dengan jabatan Kepala (bukan Wakil Kepala)
    penandatangan = (
        ASN.objects.filter(jabatan__icontains="kepala")
        .exclude(jabatan__icontains="wakil")
        .first()
    )
    if not penandatangan:
        penandatangan = ASN(nama="[Nama Kepala Sekolah]", nip="[NIP Kepala Sekolah]", pangkat="[Pangkat]", golongan="[Golongan]", jabatan="Kepala Sekolah")

    html_string = render_to_string('asn_app/laporan_cuti_pdf_template.html', {
        'asn': asn,
        'pages': pages,
        'processed_surat_cuti_list': processed_surat_cuti_list,
        'sisa_cuti': sisa_cuti_obj,
        'initial_total_sisa_cuti': initial_total_sisa_cuti,
        'penandatangan': penandatangan,
        'request': request,
        'years': years,  # Pass all selected years
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing Laporan Cuti PDF: {e}")
        return HttpResponse(f"Error writing Laporan Cuti PDF: {e}", status=500)

    safe_name = "".join(c for c in asn.nama if c.isalnum() or c in (' ', '-', '_')).rstrip()
    years_str = '_'.join(str(y) for y in years)
    filename = f'laporan_cuti_{safe_name}_{years_str}.pdf'

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


# Siswa Keluar Views
def siswa_keluar_list(request):
    """Menampilkan daftar semua siswa keluar/DO"""
    search_query = request.GET.get('q', '')

    siswa_keluar_list = SiswaKeluar.objects.all().order_by('-tanggal_keluar')

    if search_query:
        siswa_keluar_list = siswa_keluar_list.filter(
            Q(siswa__nama__icontains=search_query) |
            Q(siswa__nis__icontains=search_query) |
            Q(siswa__kelas__icontains=search_query) |
            Q(siswa__jurusan__icontains=search_query)
        )

    paginator = Paginator(siswa_keluar_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    total_siswa_keluar = SiswaKeluar.objects.count()
    total_siswa = Siswa.objects.count()
    sis_aktif = total_siswa - total_siswa_keluar

    # Calculate percentage
    if total_siswa > 0:
        percentage = (total_siswa_keluar / total_siswa) * 100
    else:
        percentage = 0

    context = {
        'page_obj': page_obj,
        'total_siswa_keluar': total_siswa_keluar,
        'total_siswa': total_siswa,
        'sis_aktif': sis_aktif,
        'percentage': percentage,
        'search_query': search_query,
    }
    return render(request, 'asn_app/siswa_keluar_list.html', context)

def siswa_keluar_detail(request, pk):
    """Menampilkan detail siswa keluar"""
    siswa_keluar = get_object_or_404(SiswaKeluar, pk=pk)
    return render(request, 'asn_app/siswa_keluar_detail.html', {'siswa_keluar': siswa_keluar})

def siswa_keluar_create(request):
    """Membuat catatan siswa keluar/DO baru"""
    if request.method == 'POST':
        form = SiswaKeluarForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data siswa keluar berhasil ditambahkan.')
            return redirect('siswa_keluar_list')
    else:
        form = SiswaKeluarForm()
    return render(request, 'asn_app/siswa_keluar_form.html', {'form': form, 'title': 'Tambah Siswa Keluar'})

def siswa_keluar_update(request, pk):
    """Mengupdate data siswa keluar"""
    siswa_keluar = get_object_or_404(SiswaKeluar, pk=pk)
    if request.method == 'POST':
        form = SiswaKeluarForm(request.POST, instance=siswa_keluar)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data siswa keluar berhasil diupdate.')
            return redirect('siswa_keluar_detail', pk=siswa_keluar.pk)
    else:
        form = SiswaKeluarForm(instance=siswa_keluar)
    return render(request, 'asn_app/siswa_keluar_form.html', {'form': form, 'title': 'Edit Siswa Keluar'})

def siswa_keluar_delete(request, pk):
    """Menghapus data siswa keluar"""
    siswa_keluar = get_object_or_404(SiswaKeluar, pk=pk)
    if request.method == 'POST':
        siswa_keluar.delete()
        messages.success(request, 'Data siswa keluar berhasil dihapus.')
        return redirect('siswa_keluar_list')
    return render(request, 'asn_app/siswa_keluar_confirm_delete.html', {'siswa_keluar': siswa_keluar})

# Siswa Views
def siswa_list(request):
    """Menampilkan daftar semua Siswa"""
    query = request.GET.get('q')
    filter_jurusan = request.GET.get('jurusan')

    if query:
        siswa_queryset = Siswa.objects.filter(nama__icontains=query).order_by('nama')
    elif filter_jurusan:
        siswa_queryset = Siswa.objects.filter(jurusan=filter_jurusan).order_by('jurusan', 'nama')
    else:
        siswa_queryset = Siswa.objects.all().order_by('nama')

    # Dashboard data
    total_siswa = Siswa.objects.count()

    # Counts by combined kelas_jurusan
    siswa_by_kelas_jurusan = Siswa.objects.filter(
        kelas__isnull=False,
        jurusan__isnull=False
    ).exclude(
        kelas__exact=''
    ).exclude(
        jurusan__exact=''
    ).values(
        'kelas', 'jurusan'
    ).annotate(
        count=Count('id')
    ).order_by('kelas', 'jurusan')


    paginator = Paginator(siswa_queryset, 10)  # 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'query': query,
        'filter_jurusan': filter_jurusan,
        'total_siswa': total_siswa,
        'siswa_by_kelas_jurusan': siswa_by_kelas_jurusan,
    }

    return render(request, 'asn_app/siswa_list.html', context)

def siswa_detail(request, pk):
    """Menampilkan detail Siswa"""
    siswa = get_object_or_404(Siswa, pk=pk)
    return render(request, 'asn_app/siswa_detail.html', {'siswa': siswa})

def siswa_create(request):
    """Membuat Siswa baru"""
    if request.method == 'POST':
        form = SiswaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('siswa_list')
    else:
        form = SiswaForm()
    return render(request, 'asn_app/siswa_form.html', {'form': form, 'title': 'Tambah Siswa'})

def siswa_update(request, pk):
    """Mengupdate data Siswa"""
    siswa = get_object_or_404(Siswa, pk=pk)
    if request.method == 'POST':
        form = SiswaForm(request.POST, request.FILES, instance=siswa)
        if form.is_valid():
            form.save()
            return redirect('siswa_detail', pk=siswa.pk)
    else:
        form = SiswaForm(instance=siswa)
    return render(request, 'asn_app/siswa_form.html', {'form': form, 'title': 'Edit Siswa'})

def siswa_delete(request, pk):
    """Menghapus data Siswa"""
    siswa = get_object_or_404(Siswa, pk=pk)
    if request.method == 'POST':
        siswa.delete()
        return redirect('siswa_list')
    return render(request, 'asn_app/siswa_confirm_delete.html', {'siswa': siswa})

def siswa_delete_all(request):
    """Menghapus semua data Siswa"""
    if request.method == 'POST':
        count, _ = Siswa.objects.all().delete()
        messages.success(request, f'{count} data Siswa berhasil dihapus.')
        return redirect('siswa_list')
    return redirect('siswa_list') # Redirect to list if GET request

def siswa_kelas_promosi(request):
    if request.method == 'POST':
        promosi_type = request.POST.get('promosi_type')
        if promosi_type == 'ix_to_x':
            qs = Siswa.objects.filter(
                Q(kelas__iexact='IX') | Q(kelas__istartswith='IX ')
            )
            updated_count = 0
            for siswa in qs:
                siswa.kelas = siswa.kelas.replace('IX', 'X', 1)
                siswa.status = 'Aktif'
                siswa.save()
                updated_count += 1
            messages.success(request, f'{updated_count} siswa berhasil dipromosikan dari Kelas IX ke Kelas X.')
        elif promosi_type == 'xii_to_lulus':
            qs = Siswa.objects.filter(
                Q(kelas__iexact='XII') | Q(kelas__istartswith='XII ')
            )
            updated_count = 0
            for siswa in qs:
                siswa.status = 'Lulus'
                siswa.save()
                updated_count += 1
            messages.success(request, f'{updated_count} siswa Kelas XII berhasil ditandai sebagai Lulus.')
        elif promosi_type == 'x_to_xii':
            qs = Siswa.objects.filter(
                Q(kelas__iexact='X') | Q(kelas__istartswith='X ')
            ).exclude(
                Q(kelas__istartswith='XI') | Q(kelas__istartswith='XII')
            )
            updated_count = 0
            for siswa in qs:
                siswa.kelas = siswa.kelas.replace('X', 'XII', 1)
                siswa.status = 'Aktif'
                siswa.save()
                updated_count += 1
            messages.success(request, f'{updated_count} siswa berhasil dipromosikan dari Kelas X ke Kelas XII.')
        return redirect('siswa_list')
    
    context = {
        'title': 'Promosi Kelas Siswa',
    }
    return render(request, 'asn_app/siswa_kelas_promosi.html', context)

# Surat Keterangan Views
def surat_keterangan_list(request):
    logger = logging.getLogger(__name__)
    queryset = SuratKeterangan.objects.filter(pk__isnull=False).order_by('-created_at')
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'asn_app/surat_keterangan_list.html', {'page_obj': page_obj})

def surat_keterangan_detail(request, pk):
    surat = get_object_or_404(SuratKeterangan, pk=pk)
    return render(request, 'asn_app/surat_keterangan_detail.html', {'surat': surat})

def surat_keterangan_create(request):
    logger = logging.getLogger(__name__)
    if request.method == 'POST':
        logger.info(f"Surat Keterangan POST data: {request.POST}")
        form = SuratKeteranganForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('surat_keterangan_list')
        else:
            logger.error(f"Surat Keterangan form errors: {form.errors.as_json()}")
    else:
        form = SuratKeteranganForm()
    return render(request, 'asn_app/surat_keterangan_form.html', {'form': form, 'title': 'Tambah Surat Keterangan'})

def surat_keterangan_update(request, pk):
    surat = get_object_or_404(SuratKeterangan, pk=pk)
    if request.method == 'POST':
        form = SuratKeteranganForm(request.POST, instance=surat)
        if form.is_valid():
            form.save()
            return redirect('surat_keterangan_detail', pk=surat.pk)
    else:
        form = SuratKeteranganForm(instance=surat)
    return render(request, 'asn_app/surat_keterangan_form.html', {'form': form, 'title': 'Edit Surat Keterangan'})

def surat_keterangan_delete(request, pk):
    surat = get_object_or_404(SuratKeterangan, pk=pk)
    if request.method == 'POST':
        surat.delete()
        return redirect('surat_keterangan_list')
    return render(request, 'asn_app/surat_keterangan_confirm_delete.html', {'surat': surat})

def surat_keterangan_export_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    surat = get_object_or_404(SuratKeterangan, pk=pk)

    kop_surat_base64 = None
    if surat.kop_surat and surat.kop_surat.gambar:
        if os.path.exists(surat.kop_surat.gambar.path):
            try:
                with open(surat.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64: {e}")

    html_string = render_to_string('asn_app/surat_keterangan_pdf_template.html', {
        'surat': surat,
        'request': request,
        'kop_surat_base64': kop_surat_base64,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=surat_keterangan_{surat.nomor_surat}.pdf'

    return response

# Surat Rekomendasi Studi Lanjut Views
def surat_rekomendasi_list(request):
    logger = logging.getLogger(__name__)
    queryset = SuratRekomendasiStudiLanjut.objects.filter(pk__isnull=False).order_by('-created_at')
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'asn_app/surat_rekomendasi_list.html', {'page_obj': page_obj})

def surat_rekomendasi_detail(request, pk):
    surat = get_object_or_404(SuratRekomendasiStudiLanjut, pk=pk)
    return render(request, 'asn_app/surat_rekomendasi_detail.html', {'surat': surat})

def surat_rekomendasi_create(request):
    logger = logging.getLogger(__name__)
    if request.method == 'POST':
        logger.info(f"Surat Rekomendasi POST data: {request.POST}")
        form = SuratRekomendasiStudiLanjutForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('surat_rekomendasi_list')
        else:
            logger.error(f"Surat Rekomendasi form errors: {form.errors.as_json()}")
    else:
        form = SuratRekomendasiStudiLanjutForm()
    return render(request, 'asn_app/surat_rekomendasi_form.html', {'form': form, 'title': 'Tambah Surat Rekomendasi Studi Lanjut'})

def surat_rekomendasi_update(request, pk):
    surat = get_object_or_404(SuratRekomendasiStudiLanjut, pk=pk)
    if request.method == 'POST':
        form = SuratRekomendasiStudiLanjutForm(request.POST, instance=surat)
        if form.is_valid():
            form.save()
            return redirect('surat_rekomendasi_detail', pk=surat.pk)
    else:
        form = SuratRekomendasiStudiLanjutForm(instance=surat)
    return render(request, 'asn_app/surat_rekomendasi_form.html', {'form': form, 'title': 'Edit Surat Rekomendasi Studi Lanjut'})

def surat_rekomendasi_delete(request, pk):
    surat = get_object_or_404(SuratRekomendasiStudiLanjut, pk=pk)
    if request.method == 'POST':
        surat.delete()
        return redirect('surat_rekomendasi_list')
    return render(request, 'asn_app/surat_rekomendasi_confirm_delete.html', {'surat': surat})

def surat_rekomendasi_export_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    surat = get_object_or_404(SuratRekomendasiStudiLanjut, pk=pk)

    kop_surat_base64 = None
    if surat.kop_surat and surat.kop_surat.gambar:
        if os.path.exists(surat.kop_surat.gambar.path):
            try:
                with open(surat.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64: {e}")

    html_string = render_to_string('asn_app/surat_rekomendasi_pdf_template.html', {
        'surat': surat,
        'request': request,
        'kop_surat_base64': kop_surat_base64,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=surat_rekomendasi_{surat.nomor_surat}.pdf'

    return response

# Surat KP4 Views
def surat_kp4_list(request):
    surat_list = SuratKP4.objects.all().order_by('-created_at')
    return render(request, 'asn_app/surat_kp4_list.html', {'surat_list': surat_list})

def surat_kp4_detail(request, pk):
    surat = get_object_or_404(SuratKP4, pk=pk)
    return render(request, 'asn_app/surat_kp4_detail.html', {'surat': surat})

def surat_kp4_create(request):
    if request.method == 'POST':
        form = SuratKP4Form(request.POST)
        formset = AnggotaKeluargaKP4FormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            surat = form.save()
            formset.instance = surat
            formset.save()
            return redirect('surat_kp4_detail', pk=surat.pk)
    else:
        form = SuratKP4Form()
        formset = AnggotaKeluargaKP4FormSet()
    return render(request, 'asn_app/surat_kp4_form.html', {
        'form': form, 'formset': formset, 'title': 'Tambah Surat KP4'
    })

def surat_kp4_update(request, pk):
    surat = get_object_or_404(SuratKP4, pk=pk)
    if request.method == 'POST':
        form = SuratKP4Form(request.POST, instance=surat)
        formset = AnggotaKeluargaKP4FormSet(request.POST, instance=surat)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('surat_kp4_detail', pk=surat.pk)
    else:
        form = SuratKP4Form(instance=surat)
        formset = AnggotaKeluargaKP4FormSet(instance=surat)
    return render(request, 'asn_app/surat_kp4_form.html', {
        'form': form, 'formset': formset, 'title': 'Edit Surat KP4'
    })

def surat_kp4_delete(request, pk):
    surat = get_object_or_404(SuratKP4, pk=pk)
    if request.method == 'POST':
        surat.delete()
        return redirect('surat_kp4_list')
    return render(request, 'asn_app/surat_kp4_confirm_delete.html', {'surat': surat})

def surat_kp4_export_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    surat = get_object_or_404(SuratKP4, pk=pk)

    kop_surat_base64 = None
    if surat.kop_surat and surat.kop_surat.gambar:
        if os.path.exists(surat.kop_surat.gambar.path):
            try:
                with open(surat.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64: {e}")

    html_string = render_to_string('asn_app/surat_kp4_pdf_template.html', {
        'surat': surat,
        'request': request,
        'kop_surat_base64': kop_surat_base64,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=kp4_{surat.pk}.pdf'

    return response

# Surat Resmi Views
def surat_resmi_list(request):
    surat_list = SuratResmi.objects.all().order_by('-created_at')
    return render(request, 'asn_app/surat_resmi_list.html', {'surat_list': surat_list})

def surat_resmi_detail(request, pk):
    surat = get_object_or_404(SuratResmi, pk=pk)
    return render(request, 'asn_app/surat_resmi_detail.html', {'surat': surat})

def surat_resmi_create(request):
    if request.method == 'POST':
        form = SuratResmiForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('surat_resmi_list')
    else:
        form = SuratResmiForm()
    return render(request, 'asn_app/surat_resmi_form.html', {'form': form, 'title': 'Tambah Surat Resmi'})

def surat_resmi_update(request, pk):
    surat = get_object_or_404(SuratResmi, pk=pk)
    if request.method == 'POST':
        form = SuratResmiForm(request.POST, instance=surat)
        if form.is_valid():
            form.save()
            return redirect('surat_resmi_detail', pk=surat.pk)
    else:
        form = SuratResmiForm(instance=surat)
    return render(request, 'asn_app/surat_resmi_form.html', {'form': form, 'title': 'Edit Surat Resmi'})

def surat_resmi_delete(request, pk):
    surat = get_object_or_404(SuratResmi, pk=pk)
    if request.method == 'POST':
        surat.delete()
        return redirect('surat_resmi_list')
    return render(request, 'asn_app/surat_resmi_confirm_delete.html', {'surat': surat})

def surat_resmi_export_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML

    surat = get_object_or_404(SuratResmi, pk=pk)

    kop_surat_base64 = None
    if surat.kop_surat and surat.kop_surat.gambar:
        if os.path.exists(surat.kop_surat.gambar.path):
            try:
                with open(surat.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64: {e}")

    html_string = render_to_string('asn_app/surat_resmi_pdf_template.html', {
        'surat': surat,
        'request': request,
        'kop_surat_base64': kop_surat_base64,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=surat_resmi_{surat.nomor}.pdf'

    return response

def export_siswa_excel(request):
    """Export all Siswa data to an Excel file."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    siswas = Siswa.objects.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Siswa Data'

    # Write headers
    headers = [
        'NIS', 'Nama', 'Kelas', 'Jurusan', 'Alamat', 'No HP', 'Nama Orang Tua'
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Write data
    for row_num, siswa in enumerate(siswas, 2):
        sheet[f'A{row_num}'] = siswa.nis
        sheet[f'B{row_num}'] = siswa.nama
        sheet[f'C{row_num}'] = siswa.kelas
        sheet[f'D{row_num}'] = siswa.jurusan
        sheet[f'E{row_num}'] = siswa.alamat
        sheet[f'F{row_num}'] = siswa.no_hp
        sheet[f'G{row_num}'] = siswa.nama_orang_tua

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=siswa_data.xlsx'
    workbook.save(response)
    return response


def import_siswa_excel(request):
    """Import Siswa data from an Excel file."""
    if request.method == 'POST':
        import openpyxl
        excel_file = request.FILES['excel_file']
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            (
                nama, nis, kelas, jurusan, alamat, no_hp, nama_orang_tua
            ) = row # Assuming these are the columns in the excel

            Siswa.objects.create(
                nama=nama,
                nis=nis,
                kelas=kelas,
                jurusan=jurusan,
                alamat=alamat,
                no_hp=no_hp,
                nama_orang_tua=nama_orang_tua
            )
        return redirect('siswa_list')
    return render(request, 'asn_app/import_form.html')


def export_siswa_keluar_excel(request):
    """Export all SiswaKeluar data to an Excel file."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    siswa_keluar_list = SiswaKeluar.objects.all().order_by('-tanggal_keluar')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Siswa Keluar Data'

    # Write headers
    headers = [
        'No', 'Nama Siswa', 'NIS', 'Kelas', 'Jurusan',
        'Tanggal Keluar', 'Alasan Keluar', 'Tanggal Dicatat'
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Write data
    for row_num, siswa_keluar in enumerate(siswa_keluar_list, 2):
        sheet[f'A{row_num}'] = row_num - 1
        sheet[f'B{row_num}'] = siswa_keluar.siswa.nama
        sheet[f'C{row_num}'] = siswa_keluar.siswa.nis
        sheet[f'D{row_num}'] = siswa_keluar.siswa.kelas
        sheet[f'E{row_num}'] = siswa_keluar.siswa.jurusan
        sheet[f'F{row_num}'] = siswa_keluar.tanggal_keluar
        sheet[f'G{row_num}'] = siswa_keluar.alasan_keluar
        sheet[f'H{row_num}'] = siswa_keluar.created_at.strftime('%Y-%m-%d %H:%M:%S')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=siswa_keluar_data.xlsx'
    workbook.save(response)
    return response


def export_siswa_keluar_pdf(request):
    """Export all SiswaKeluar data to a PDF file."""
    from weasyprint import HTML
    import os

    siswa_keluar_list = SiswaKeluar.objects.all().order_by('-tanggal_keluar')
    total_siswa_keluar = siswa_keluar_list.count()
    total_siswa = Siswa.objects.count()
    sis_aktif = total_siswa - total_siswa_keluar

    html_string = render_to_string('asn_app/siswa_keluar_pdf_template.html', {
        'siswa_keluar_list': siswa_keluar_list,
        'total_siswa_keluar': total_siswa_keluar,
        'total_siswa': total_siswa,
        'sis_aktif': sis_aktif,
        'generated_at': now(),
    })

    # Create PDF
    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    # Create HTTP response with PDF
    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=siswa_keluar_data.pdf'
    return response



class SuratResmiListView(ListView):
    model = SuratResmi
    template_name = 'asn_app/surat_resmi_list.html'
    context_object_name = 'surat_resmi_list'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(nomor__icontains=query) |
                Q(perihal__icontains=query) |
                Q(pejabat_tujuan_surat__icontains=query)
            )
        return queryset.order_by('-created_at')


class SuratResmiDetailView(DetailView):
    model = SuratResmi
    template_name = 'asn_app/surat_resmi_detail.html'
    context_object_name = 'surat_resmi'


class SuratResmiCreateView(CreateView):
    model = SuratResmi
    form_class = SuratResmiForm
    template_name = 'asn_app/surat_resmi_form.html'

    def get_success_url(self):
        messages.success(self.request, 'Surat Resmi berhasil dibuat.')
        return reverse_lazy('surat_resmi_detail', kwargs={'pk': self.object.pk})


class SuratResmiUpdateView(UpdateView):
    model = SuratResmi
    form_class = SuratResmiForm
    template_name = 'asn_app/surat_resmi_form.html'

    def get_success_url(self):
        messages.success(self.request, 'Surat Resmi berhasil diperbarui.')
        return reverse_lazy('surat_resmi_detail', kwargs={'pk': self.object.pk})


class SuratResmiDeleteView(DeleteView):
    model = SuratResmi
    template_name = 'asn_app/surat_resmi_confirm_delete.html'
    success_url = reverse_lazy('surat_resmi_list')
    context_object_name = 'surat_resmi'

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Surat Resmi berhasil dihapus.')
        return super().delete(request, *args, **kwargs)


# SPTJM Views
class SPTJMListView(ListView):
    model = SPTJM
    template_name = 'asn_app/sptjm_list.html'
    context_object_name = 'sptjm_list'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(nomor_surat__icontains=query) |
                Q(isi_surat__icontains=query)
            )
        return queryset.order_by('-created_at')


class SPTJMDetailView(DetailView):
    model = SPTJM
    template_name = 'asn_app/sptjm_detail.html'
    context_object_name = 'sptjm'


class SPTJMCreateView(CreateView):
    model = SPTJM
    form_class = SPTJMForm
    template_name = 'asn_app/sptjm_form.html'

    def get_success_url(self):
        messages.success(self.request, 'SPTJM berhasil dibuat.')
        return reverse_lazy('sptjm_detail', kwargs={'pk': self.object.pk})


class SPTJMUpdateView(UpdateView):
    model = SPTJM
    form_class = SPTJMForm
    template_name = 'asn_app/sptjm_form.html'

    def get_success_url(self):
        messages.success(self.request, 'SPTJM berhasil diperbarui.')
        return reverse_lazy('sptjm_detail', kwargs={'pk': self.object.pk})


class SPTJMDeleteView(DeleteView):
    model = SPTJM
    template_name = 'asn_app/sptjm_confirm_delete.html'
    success_url = reverse_lazy('sptjm_list')
    context_object_name = 'sptjm'

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'SPTJM berhasil dihapus.')
        return super().delete(request, *args, **kwargs)


def sptjm_export_pdf(request, pk):
    import os
    import base64
    import logging # Import logging
    from weasyprint import HTML, CSS

    logger = logging.getLogger(__name__) # Get logger instance

    sptjm = get_object_or_404(SPTJM, pk=pk);

    kop_surat_base64 = None
    if sptjm.kop_surat and sptjm.kop_surat.gambar:
        logger.info(f"Kop surat found: {sptjm.kop_surat.nama}, image path: {sptjm.kop_surat.gambar.path}")
        if os.path.exists(sptjm.kop_surat.gambar.path):
            try:
                with open(sptjm.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = sptjm.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
                    logger.info("Kop surat image encoded to Base64 successfully.")
            except Exception as e:
                logger.error(f"Error encoding image to base64 for SPTJM PDF: {e}")
        else:
            logger.warning(f"Kop surat image file not found at: {sptjm.kop_surat.gambar.path}")
    else:
        logger.info("No kop surat or image found for this SPTJM.")

    html_string = render_to_string('asn_app/sptjm_pdf_template.html', {
        'sptjm': sptjm,
        'kop_surat_base64': kop_surat_base64,
    })

    # Adding CSS for layout as requested
    css_string = """
        @page {
            size: A4;
            margin: 2.5cm;
        }
        body {
            font-family: 'Times New Roman', serif;
            font-size: 12pt;
            line-height: 1.5;
        }
        .text-center {
            text-align: center;
        }
        .text-left {
            text-align: left;
        }
        .signer-section {
            padding-left: 250pt;
            text-align: left;
        }
        .signer-details {
            margin-top: 50pt;
        }
        .employee-list {
            list-style-type: decimal;
            padding-left: 40px;
        }
        .details-table {
            border-collapse: collapse;
            width: 100%;
        }
        .details-table td {
            padding: 2px 0;
            vertical-align: top;
        }
        .details-table .label {
            width: 120px;
        }
        .details-table .colon {
            width: 10px;
        }
        /* Kop Surat specific styles - already in template, but ensuring consistency */
        .kop-surat {
            text-align: center;
            margin-bottom: 20px;
            border-bottom: 2px solid #000;
            padding-bottom: 10px;
        }
        .kop-surat img {
            max-width: 100%;
            height: auto;
        }
    """

    try:
        # For Base64 images, often base_url=None works better as it prevents WeasyPrint
        # from trying to resolve the base64 string as a relative URL.
        html = HTML(string=html_string, base_url=None)
        result = html.write_pdf(stylesheets=[CSS(string=css_string)])
        logger.info("SPTJM PDF generated successfully.")
    except Exception as e:
        logger.error(f"Error writing SPTJM PDF: {e}", exc_info=True)
        return HttpResponse(f"Error writing SPTJM PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=sptjm_{sptjm.nomor_surat}.pdf'

    return response

# SPMT Views
class SPMTListView(ListView):
    model = SPMT
    template_name = 'asn_app/spmt_list.html'
    context_object_name = 'spmt_list'
    paginate_by = 10

class SPMTDetailView(DetailView):
    model = SPMT
    template_name = 'asn_app/spmt_detail.html'
    context_object_name = 'spmt'

class SPMTCreateView(CreateView):
    model = SPMT
    form_class = SPMTForm
    template_name = 'asn_app/spmt_form.html'
    success_url = reverse_lazy('spmt_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Tambah SPMT'
        return context

class SPMTUpdateView(UpdateView):
    model = SPMT
    form_class = SPMTForm
    template_name = 'asn_app/spmt_form.html'
    context_object_name = 'spmt'

    def get_success_url(self):
        return reverse_lazy('spmt_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit SPMT'
        return context

class SPMTDeleteView(DeleteView):
    model = SPMT
    template_name = 'asn_app/spmt_confirm_delete.html'
    success_url = reverse_lazy('spmt_list')
    context_object_name = 'spmt'

def spmt_export_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML, CSS

    spmt = get_object_or_404(SPMT, pk=pk)

    kop_surat_base64 = None
    if spmt.kop_surat and spmt.kop_surat.gambar:
        if os.path.exists(spmt.kop_surat.gambar.path):
            try:
                with open(spmt.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = spmt.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64: {e}")

    html_string = render_to_string('asn_app/spmt_pdf_template.html', {
        'spmt': spmt,
        'kop_surat_base64': kop_surat_base64,
    })

    css_string = """
        body { font-family: 'Times New Roman', serif; font-size: 12pt; }
        .center { text-align: center; }
        .left { text-align: left; }
        .justify { text-align: justify; }
        .signature { padding-left: 250pt; }
        table { border-collapse: collapse; width: 100%; }
        td { vertical-align: top; }
        .label { width: 150px; }
        .colon { width: 10px; }
        .kop-surat img { max-width: 100%; height: auto; }
    """

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf(stylesheets=[CSS(string=css_string)])
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=SPMT_{spmt.nomor_surat}.pdf'

    return response

def spmt_export_excel(request):
    """Export all SPMT data to an Excel file."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    spmts = SPMT.objects.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'SPMT Data'

    # Write headers
    headers = [
        'Nomor Surat', 'Tempat Ditetapkan', 'Tanggal Surat',
        'Penandatangan (Nama)', 'Penandatangan (NIP)', 'Pegawai (Nama)', 'Pegawai (NIP)',
        'Peraturan', 'Nomor Peraturan', 'Tahun Peraturan', 'Tentang',
        'Tanggal Terhitung Mulai', 'Sebagai', 'Tempat Tugas',
        'Created At', 'Updated At'
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Write data
    for row_num, spmt in enumerate(spmts, 2):
        penandatangan_nama = spmt.penandatangan.nama if spmt.penandatangan else ''
        penandatangan_nip = spmt.penandatangan.nip if spmt.penandatangan else ''
        pegawai_nama = spmt.pegawai.nama if spmt.pegawai else ''
        pegawai_nip = spmt.pegawai.nip if spmt.pegawai else ''

        sheet[f'A{row_num}'] = spmt.nomor_surat
        sheet[f'B{row_num}'] = spmt.tempat_ditetapkan
        sheet[f'C{row_num}'] = spmt.tanggal_surat.strftime('%Y-%m-%d') if spmt.tanggal_surat else ''
        sheet[f'D{row_num}'] = penandatangan_nama
        sheet[f'E{row_num}'] = penandatangan_nip
        sheet[f'F{row_num}'] = pegawai_nama
        sheet[f'G{row_num}'] = pegawai_nip
        sheet[f'H{row_num}'] = spmt.peraturan
        sheet[f'I{row_num}'] = spmt.nomor_peraturan
        sheet[f'J{row_num}'] = spmt.tahun_peraturan
        sheet[f'K{row_num}'] = spmt.tentang
        sheet[f'L{row_num}'] = spmt.tanggal_terhitung.strftime('%Y-%m-%d') if spmt.tanggal_terhitung else ''
        sheet[f'M{row_num}'] = spmt.sebagai
        sheet[f'N{row_num}'] = spmt.tempat_tugas
        sheet[f'O{row_num}'] = spmt.created_at.strftime('%Y-%m-%d %H:%M:%S') if spmt.created_at else ''
        sheet[f'P{row_num}'] = spmt.updated_at.strftime('%Y-%m-%d %H:%M:%S') if spmt.updated_at else ''

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=spmt_data.xlsx'
    workbook.save(response)
    return response


def spmt_import_excel(request):
    """Import SPMT data from an Excel file."""
    if request.method == 'POST':
        import openpyxl
        from datetime import datetime

        excel_file = request.FILES['excel_file']
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        success_count = 0
        error_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                (
                    nomor_surat, tempat_ditetapkan, tanggal_surat_val,
                    penandatangan_nama, penandatangan_nip, pegawai_nama, pegawai_nip,
                    peraturan, nomor_peraturan, tahun_peraturan, tentang,
                    tanggal_terhitung_val, sebagai, tempat_tugas
                ) = row

                # Handle date conversion
                tanggal_surat = None
                if isinstance(tanggal_surat_val, datetime):
                    tanggal_surat = tanggal_surat_val.date()
                elif isinstance(tanggal_surat_val, str):
                    try:
                        tanggal_surat = datetime.strptime(tanggal_surat_val, '%Y-%m-%d').date()
                    except ValueError:
                        pass

                tanggal_terhitung = None
                if isinstance(tanggal_terhitung_val, datetime):
                    tanggal_terhitung = tanggal_terhitung_val.date()
                elif isinstance(tanggal_terhitung_val, str):
                    try:
                        tanggal_terhitung = datetime.strptime(tanggal_terhitung_val, '%Y-%m-%d').date()
                    except ValueError:
                        pass

                # Find penandatangan and pegawai by name (or NIP if available)
                penandatangan = None
                if penandatangan_nama:
                    penandatangan = ASN.objects.filter(nama=penandatangan_nama).first()
                    if not penandatangan and penandatangan_nip:
                        penandatangan = ASN.objects.filter(nip=penandatangan_nip).first()

                pegawai = None
                if pegawai_nama:
                    pegawai = ASN.objects.filter(nama=pegawai_nama).first()
                    if not pegawai and pegawai_nip:
                        pegawai = ASN.objects.filter(nip=pegawai_nip).first()

                SPMT.objects.create(
                    nomor_surat=nomor_surat,
                    tempat_ditetapkan=tempat_ditetapkan,
                    tanggal_surat=tanggal_surat,
                    penandatangan=penandatangan,
                    pegawai=pegawai,
                    peraturan=peraturan,
                    nomor_peraturan=nomor_peraturan,
                    tahun_peraturan=str(tahun_peraturan) if tahun_peraturan else '',
                    tentang=tentang,
                    tanggal_terhitung=tanggal_terhitung,
                    sebagai=sebagai,
                    tempat_tugas=tempat_tugas
                )
                success_count += 1
            except Exception as e:
                logging.error(f"Error importing SPMT row: {e}")
                error_count += 1

        messages.success(request, f'Successfully imported {success_count} SPMT records. {error_count} errors.')
        return redirect('spmt_list')

    return render(request, 'asn_app/spmt_import_form.html')


# Surat Umum Views
class SuratUmumListView(ListView):
    model = SuratUmum
    template_name = 'asn_app/surat_umum_list.html'
    context_object_name = 'surat_umum_list'
    paginate_by = 10

class SuratUmumDetailView(DetailView):
    model = SuratUmum
    template_name = 'asn_app/surat_umum_detail.html'
    context_object_name = 'surat'

class SuratUmumCreateView(CreateView):
    model = SuratUmum
    form_class = SuratUmumForm
    template_name = 'asn_app/surat_umum_form.html'
    success_url = reverse_lazy('surat_umum_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Tambah Surat Umum'
        return context

class SuratUmumUpdateView(UpdateView):
    model = SuratUmum
    form_class = SuratUmumForm
    template_name = 'asn_app/surat_umum_form.html'
    context_object_name = 'surat'

    def get_success_url(self):
        return reverse_lazy('surat_umum_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit Surat Umum'
        return context

class SuratUmumDeleteView(DeleteView):
    model = SuratUmum
    template_name = 'asn_app/surat_umum_confirm_delete.html'
    success_url = reverse_lazy('surat_umum_list')
    context_object_name = 'surat'

def surat_umum_export_pdf(request, pk):
    import os
    import base64
    from weasyprint import HTML, CSS

    surat = get_object_or_404(SuratUmum, pk=pk)

    kop_surat_base64 = None
    if surat.kop_surat and surat.kop_surat.gambar:
        if os.path.exists(surat.kop_surat.gambar.path):
            try:
                with open(surat.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64: {e}")

    html_string = render_to_string('asn_app/surat_umum_pdf_template.html', {
        'surat': surat,
        'kop_surat_base64': kop_surat_base64,
    })

    css_string = """
        body { font-family: 'Times New Roman', serif; font-size: 12pt; }
        .center { text-align: center; }
        .left { text-align: left; }
        .justify { text-align: justify; }
        .signature { padding-left: 250pt; }
    """

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf(stylesheets=[CSS(string=css_string)])
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=surat_umum_{surat.pk}.pdf'

    return response


# Surat Panggilan Siswa Views
def surat_panggilan_siswa_list(request):
    """Menampilkan daftar semua Surat Panggilan Siswa"""
    surat_list = SuratPanggilanSiswa.objects.all().order_by('-created_at')
    return render(request, 'asn_app/surat_panggilan_siswa_list.html', {'surat_list': surat_list})

def surat_panggilan_siswa_detail(request, pk):
    """Menampilkan detail Surat Panggilan Siswa"""
    surat = get_object_or_404(SuratPanggilanSiswa, pk=pk)
    return render(request, 'asn_app/surat_panggilan_siswa_detail.html', {'surat': surat})

def surat_panggilan_siswa_create(request):
    """Membuat Surat Panggilan Siswa baru"""
    if request.method == 'POST':
        form = SuratPanggilanSiswaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('surat_panggilan_siswa_list')
    else:
        form = SuratPanggilanSiswaForm()
    return render(request, 'asn_app/surat_panggilan_siswa_form.html', {'form': form, 'title': 'Tambah Surat Panggilan Siswa'})

def surat_panggilan_siswa_update(request, pk):
    """Mengupdate Surat Panggilan Siswa"""
    surat = get_object_or_404(SuratPanggilanSiswa, pk=pk)
    if request.method == 'POST':
        form = SuratPanggilanSiswaForm(request.POST, instance=surat)
        if form.is_valid():
            form.save()
            return redirect('surat_panggilan_siswa_detail', pk=surat.pk)
    else:
        form = SuratPanggilanSiswaForm(instance=surat)
    return render(request, 'asn_app/surat_panggilan_siswa_form.html', {'form': form, 'title': 'Edit Surat Panggilan Siswa'})

def surat_panggilan_siswa_delete(request, pk):
    """Menghapus Surat Panggilan Siswa"""
    surat = get_object_or_404(SuratPanggilanSiswa, pk=pk)
    if request.method == 'POST':
        surat.delete()
        return redirect('surat_panggilan_siswa_list')
    return render(request, 'asn_app/surat_panggilan_siswa_confirm_delete.html', {'surat': surat})

def surat_panggilan_siswa_export_pdf(request, pk):
    """Export Surat Panggilan Siswa ke PDF"""
    import os
    import base64
    from weasyprint import HTML, CSS

    surat = get_object_or_404(SuratPanggilanSiswa, pk=pk)

    kop_surat_base64 = None
    if surat.kop_surat and surat.kop_surat.gambar:
        if os.path.exists(surat.kop_surat.gambar.path):
            try:
                with open(surat.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64: {e}")

    html_string = render_to_string('asn_app/surat_panggilan_siswa_pdf_template.html', {
        'surat': surat,
        'kop_surat_base64': kop_surat_base64,
        'request': request,
    })

    css_string = """
        body { font-family: 'Times New Roman', serif; font-size: 12pt; }
        .center { text-align: center; }
        .left { text-align: left; }
        .justify { text-align: justify; }
        .signature { padding-left: 250pt; }
    """

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf(stylesheets=[CSS(string=css_string)])
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=surat_panggilan_{surat.siswa.nama}_{surat.nomor_surat}.pdf'

    return response


def surat_undangan_siswa_list(request):
    """Menampilkan daftar semua Surat Undangan Siswa"""
    surat_list = SuratUndanganSiswa.objects.all().order_by('-created_at')
    return render(request, 'asn_app/surat_undangan_siswa_list.html', {'surat_list': surat_list})

def surat_undangan_siswa_detail(request, pk):
    """Menampilkan detail Surat Undangan Siswa"""
    surat = get_object_or_404(SuratUndanganSiswa, pk=pk)
    return render(request, 'asn_app/surat_undangan_siswa_detail.html', {'surat': surat})

def surat_undangan_siswa_create(request):
    """Membuat Surat Undangan Siswa baru"""
    if request.method == 'POST':
        form = SuratUndanganSiswaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('surat_undangan_siswa_list')
    else:
        form = SuratUndanganSiswaForm()
    return render(request, 'asn_app/surat_undangan_siswa_form.html', {'form': form, 'title': 'Tambah Surat Undangan Siswa'})

def surat_undangan_siswa_update(request, pk):
    """Mengupdate Surat Undangan Siswa"""
    surat = get_object_or_404(SuratUndanganSiswa, pk=pk)
    if request.method == 'POST':
        form = SuratUndanganSiswaForm(request.POST, instance=surat)
        if form.is_valid():
            form.save()
            return redirect('surat_undangan_siswa_detail', pk=surat.pk)
    else:
        form = SuratUndanganSiswaForm(instance=surat)
    return render(request, 'asn_app/surat_undangan_siswa_form.html', {'form': form, 'title': 'Edit Surat Undangan Siswa'})

def surat_undangan_siswa_delete(request, pk):
    """Menghapus Surat Undangan Siswa"""
    surat = get_object_or_404(SuratUndanganSiswa, pk=pk)
    if request.method == 'POST':
        surat.delete()
        return redirect('surat_undangan_siswa_list')
    return render(request, 'asn_app/surat_undangan_siswa_confirm_delete.html', {'surat': surat})

def surat_undangan_siswa_export_pdf(request, pk):
    """Export Surat Undangan Siswa ke PDF"""
    import os
    import base64
    from weasyprint import HTML, CSS

    surat = get_object_or_404(SuratUndanganSiswa, pk=pk)

    kop_surat_base64 = None
    if surat.kop_surat and surat.kop_surat.gambar:
        if os.path.exists(surat.kop_surat.gambar.path):
            try:
                with open(surat.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64: {e}")

    html_string = render_to_string('asn_app/surat_undangan_siswa_pdf_template.html', {
        'surat': surat,
        'kop_surat_base64': kop_surat_base64,
        'request': request,
    })

    css_string = """
        body { font-family: 'Times New Roman', serif; font-size: 12pt; }
        .center { text-align: center; }
        .left { text-align: left; }
        .justify { text-align: justify; }
        .signature { padding-left: 250pt; }
    """

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf(stylesheets=[CSS(string=css_string)])
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    siswa_nama = surat.siswa.nama if surat.siswa else ''
    response['Content-Disposition'] = f'attachment; filename=surat_undangan_{siswa_nama}_{surat.nomor_surat}.pdf'

    return response


# Surat Dispensasi Views
def surat_dispensasi_list(request):
    """Menampilkan daftar semua Surat Dispensasi"""
    surat_list = SuratDispensasi.objects.all().order_by('-created_at')
    return render(request, 'asn_app/surat_dispensasi_list.html', {'surat_list': surat_list})


def surat_dispensasi_detail(request, pk):
    """Menampilkan detail Surat Dispensasi"""
    surat = get_object_or_404(SuratDispensasi, pk=pk)
    return render(request, 'asn_app/surat_dispensasi_detail.html', {'surat': surat})


def surat_dispensasi_create(request):
    """Membuat Surat Dispensasi baru"""
    if request.method == 'POST':
        form = SuratDispensasiForm(request.POST)
        formset = PesertaDispensasiFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            surat = form.save()
            formset.instance = surat
            formset.save()
            messages.success(request, 'Surat Dispensasi berhasil dibuat.')
            return redirect('surat_dispensasi_detail', pk=surat.pk)
        else:
            messages.error(request, 'Terdapat kesalahan pada formulir yang diisi.')
    else:
        form = SuratDispensasiForm()
        formset = PesertaDispensasiFormSet()
    return render(request, 'asn_app/surat_dispensasi_form.html', {'form': form, 'formset': formset, 'title': 'Tambah Surat Dispensasi'})


def surat_dispensasi_update(request, pk):
    """Mengupdate Surat Dispensasi"""
    surat = get_object_or_404(SuratDispensasi, pk=pk)
    if request.method == 'POST':
        form = SuratDispensasiForm(request.POST, instance=surat)
        formset = PesertaDispensasiFormSet(request.POST, instance=surat)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Surat Dispensasi berhasil diperbarui.')
            return redirect('surat_dispensasi_detail', pk=surat.pk)
        else:
            messages.error(request, 'Terdapat kesalahan pada formulir yang diisi.')
    else:
        form = SuratDispensasiForm(instance=surat)
        formset = PesertaDispensasiFormSet(instance=surat)
    return render(request, 'asn_app/surat_dispensasi_form.html', {'form': form, 'formset': formset, 'title': 'Edit Surat Dispensasi'})


def surat_dispensasi_delete(request, pk):
    """Menghapus Surat Dispensasi"""
    surat = get_object_or_404(SuratDispensasi, pk=pk)
    if request.method == 'POST':
        surat.delete()
        messages.success(request, 'Surat Dispensasi berhasil dihapus.')
        return redirect('surat_dispensasi_list')
    return render(request, 'asn_app/surat_dispensasi_confirm_delete.html', {'surat': surat})


def surat_dispensasi_export_pdf(request, pk):
    """Export Surat Dispensasi ke PDF"""
    import os
    import base64
    from weasyprint import HTML

    surat = get_object_or_404(SuratDispensasi, pk=pk)

    kop_surat_base64 = None
    if surat.kop_surat and surat.kop_surat.gambar:
        if os.path.exists(surat.kop_surat.gambar.path):
            try:
                with open(surat.kop_surat.gambar.path, 'rb') as image_file:
                    image_data = image_file.read()
                    image_format = surat.kop_surat.gambar.name.split('.')[-1].lower()
                    if image_format == 'jpg':
                        image_format = 'jpeg'
                    kop_surat_base64 = f"data:image/{image_format};base64,{base64.b64encode(image_data).decode('utf-8')}"
            except Exception as e:
                logging.error(f"Error encoding image to base64: {e}")

    html_string = render_to_string('asn_app/surat_dispensasi_pdf_template.html', {
        'surat': surat,
        'kop_surat_base64': kop_surat_base64,
        'request': request,
    })

    try:
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        result = html.write_pdf()
    except Exception as e:
        logging.error(f"Error writing PDF: {e}")
        return HttpResponse(f"Error writing PDF: {e}", status=500)

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=surat_dispensasi_{surat.nomor_surat}.pdf'

    return response
